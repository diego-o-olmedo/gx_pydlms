from gurux_dlms.manufacturersettings import GXAttributeCollection, GXDLMSAttributeSettings

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from objects_from_excel.idis.excel_checker import Checker

from gurux_dlms.GXDLMSClient import GXDLMSClient
from gurux_dlms.objects import GXDLMSObjectCollection
from gurux_dlms.enums.AccessMode import AccessMode


class ObjectsIDIS:
    # Create index and the corresponding value of the first level heading
    first_level = {
        'Type': "type",
        'Index': "index",
        'Object': "Object / Attribute Name",  # name
        'Standard': "Standard",
        'Meter': "Supported Objects",  # Supported Objects
        'Attribute': "Attribute Type",  # Attribute Type
        'Class': "Class",  # class
        'Version': "Ver.",
        'SN': "SN",
        'OBIS': "OBIS code / Default Value",
        'Access': "Access rights\n[Get, Set, Action], optional if in '( )'",
        'Comments': "Comments"
    }
    # Secondary title header, used for indexing
    title = {}

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = self.wb['Object model']
        """ Object Functions mapping """
        self.functions = {}
        self.__meter_function_mapping()
        """  Title handling """
        self.__title_handle()
        """ Self-Check """
        Checker.check(self)
        # self.objects = {}
        # self.__objects_model_creat()

    def __meter_function_mapping(self):
        sheet = self.wb['Meter Function Type']
        for row in range(2, sheet.max_row+1):
            meter_type = sheet[f"A{row}"].value
            if meter_type is not None:
                self.functions[meter_type] = []
                for col in range(2, sheet.max_column+1):
                    col = get_column_letter(col)
                    enable = sheet[f"{col}{row}"].value
                    if enable is not None:
                        self.functions[meter_type].append(sheet[f"{col}1"].value)

    def __title_handle(self):
        self.__first_level_handle()
        self.__second_level_handle()

    def __first_level_handle(self):
        for col in range(1, self.sheet.max_column+1):
            col = get_column_letter(col)
            value = self.sheet[f'{col}1'].value
            for key in self.first_level:
                if value == self.first_level[key]:
                    self.first_level[key] = col

    def __second_level_handle(self):
        for key in self.first_level:
            # Single heading, the default Comments is the last grid, and there is only one grid
            if self.sheet[f"{get_column_letter(column_index_from_string(self.first_level[key])+1)}1"].value is not None \
                    or key == 'Comments':
                self.title[key] = self.first_level[key]
            else:
                self.title[key] = {}
                for col in range(column_index_from_string(self.first_level[key]), self.sheet.max_column+1):
                    col = get_column_letter(col)
                    value = self.sheet[f'{col}2'].value
                    if col != self.first_level[key] and self.sheet[f'{col}1'].value is not None:
                        break
                    self.title[key][value] = col

    def info_print(self):
        print(f"File: {self.file_path}")
        print(f"Sheet: {self.sheet.title}")
        print(f"Title reference: {self.title}")
        print(f"All standard: {list(self.title['Standard'].keys())}")
        print(f"All meter type: {list(self.title['Meter'].keys())}")
        print(f"All client: {list(self.title['Access'].keys())}")


    @classmethod
    def get_objects(cls, file_path, meter_type, client):
        objects_src = cls(file_path)
        objects_src.info_print()

        if meter_type not in objects_src.title['Meter']:
            raise ValueError(f"Meter type error, it should in {list(cls.title['Meter'].keys())}")
        if client not in objects_src.title['Access']:
            raise ValueError(f"Client error, it should in {list(cls.title['Access'].keys())}")
        objects = GXDLMSObjectCollection()
        obj = None
        function = None
        for row in range(1, objects_src.sheet.max_row+1):
            cell_type_value = objects_src.sheet[f'{objects_src.title["Type"]}{row}'].value
            if cell_type_value == 'o':
                function = objects_src.sheet[f'{objects_src.title["Meter"][meter_type]}{row}'].value
                if function in objects_src.functions[meter_type]:
                    class_id = int(objects_src.sheet[f'{objects_src.title["Class"]}{row}'].value)
                    obj = GXDLMSClient.createObject(class_id)
                    obj.logicalName = objects_src.sheet[f'{objects_src.title["OBIS"]}{row}'].value.replace('-', '.').replace(':', '.')
                    obj.description = objects_src.sheet[f'{objects_src.title["Object"]}{row}'].value
                    obj.version = int(objects_src.sheet[f'{objects_src.title["Version"]}{row}'].value)
                    obj.comments = objects_src.sheet[f'{objects_src.title["Comments"]}{row}'].value
                    objects.append(obj)
            elif cell_type_value in ('a', 'm'):
                if function in objects_src.functions[meter_type]:
                    collection = obj.attributes if cell_type_value == 'a' else obj.methodAttributes
                    attr = GXDLMSAttributeSettings()
                    attr.index = objects_src.sheet[f'{objects_src.title["Index"]}{row}'].value
                    attr.name = objects_src.sheet[f'{objects_src.title["Object"]}{row}'].value
                    attr.type_ = objects_src.sheet[f'{objects_src.title["Attribute"]}{row}'].value
                    attr.access = AccessMode.transfer(
                        objects_src.sheet[f'{objects_src.title["Access"][client]}{row}'].value)
                    attr.values = objects_src.sheet[f'{objects_src.title["OBIS"]}{row}'].value
                    attr.comments = objects_src.sheet[f'{objects_src.title["Comments"]}{row}'].value
                    collection.append(attr)
        return objects
